import time
import traceback
import getpass
import brotli
from datetime import datetime, timedelta
from decimal import Decimal

import pandas as pd
import json
import openpyxl
from web3 import Web3
from eth_account import Account
import eth_account.messages
import itertools
import requests
from web3.middleware import geth_poa_middleware
from eth_account.messages import defunct_hash_message
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from enum import Enum
from fake_useragent import UserAgent

class Block_chain(Enum):
    ETH = ('https://mainnet.infura.io/v3/a4c5b4ef4d704823980c3349f3b99d18',1, 'https://cn.etherscan.com/tx/', 'pos', 0)
    GOERLI_TEST = ('https://rpc.ankr.com/eth_goerli', 5, 'https://goerli.etherscan.io/tx/', 'poa', 0)
    LINEA_TEST = ('https://rpc.goerli.linea.build/', 59140, 'https://explorer.goerli.linea.build/tx/', 'poa', 0)
    ZKS_ERA_TEST = ('https://testnet.era.zksync.dev', 280, 'https://zksync2-testnet.zkscan.io/tx/', 'poa', 0.25)
    ZKS_ERA = ('https://mainnet.era.zksync.io', 324, 'https://zksync2-mainnet.zkscan.io/tx/', 'pos', 0.25)
    BSC = ('https://bsc-dataseed.binance.org/', 56, 'https://bscscan.com/tx/', 'pos', 0)
    BSC_ANKR = ('https://rpc.ankr.com/bsc', 56, 'https://bscscan.com/tx/', 'pos', 1)
    BSC_TEST = ('https://endpoints.omniatech.io/v1/bsc/testnet/public/', 97, 'https://testnet.bscscan.com/tx/', 'pos', 0)
    OPTIMISM = ('https://opt-mainnet.g.alchemy.com/v2/bK9nYHNPAcdwbz1CqrbegEp6JKj0CmFe', 10, 'https://optimistic.etherscan.io/tx/', 'pos', 0)

    def __init__(self, url, chain_id, block_chain_explorer, middleware, chain_gas_price):
        self.url = url
        self.chain_id = chain_id
        self.block_chain_explorer = block_chain_explorer
        self.middleware = middleware
        self.chain_gas_price = chain_gas_price


block_chain = Block_chain.BSC_ANKR
rpc_url = block_chain.url
chainId = block_chain.chain_id
block_chain_explorer = block_chain.block_chain_explorer
chain_gas_price = block_chain.chain_gas_price


def get_time():
    return str(datetime.now())+' '

# 查询链上交互信息
def batchSerachReceiptAndSaveExcel(zzHashListAndNum):
    # key:hash,value:excel下标
    global success_count
    global need_mint
    currentSuccessCount = 0
    total_gas = Decimal('0')
    print('开始查询结果')
    count = len(zzHashListAndNum)
    i = 0
    for tx_hash in zzHashListAndNum:
        try:
            print('{0}/{1}'.format(i+1,count))
            i += 1

            receipt = w3.eth.waitForTransactionReceipt(tx_hash)
            if receipt['status'] == 1:
                use_gas = w3.fromWei(receipt['effectiveGasPrice'] * receipt['gasUsed'], 'ether')
                total_gas += use_gas
                currentSuccessCount += 1
                success_count += 1
                num = zzHashListAndNum[tx_hash]
                excel_data.loc[num:num, mint_status] = 'YES'
                excel_data.to_excel(excel_file_name, index=False, engine="openpyxl")
                print('成功 花费Gas：{0}'.format(use_gas))

            else:
                need_mint = True
                print('失败  {0}{1}'.format(block_chain_explorer,w3.toHex(tx_hash)))
        except Exception as e:
            need_mint = True
            print('发生错误：{0}'.format(e))

    print('\n本批实际Mint账号数:{0}；成功数:{1}；实际消耗gas:{2}；'.format(count, currentSuccessCount, total_gas))



# 查询链上交互信息
def searchReceipt(tx_hash):
    receipt = w3.eth.waitForTransactionReceipt(tx_hash)
    # print(get_time() + 'receipt:{0}，type{1}'.format(receipt,type(receipt)))
    return receipt


# 是否有资格领取
def get_can_claim(addr):
    # 没余额就是可以领
    balance = mintContract.functions.balanceOf(addr).call()
    return balance == 0

# file_patch：存放headers的文件路径
# replace_key要替换的key数组，定义要替换key的名字
# replace_value 要替换为的值，key和value要一一对应
def get_headers(file_patch='headers.txt',replace_key=None,replace_value=None):
    data = {}
    with open(file_patch, "r") as f:
        lines = f.readlines()
        count = len(lines)
        for i in range(count):
            line = lines[i]
            line = line.strip('\n')  # 去掉列表中每一个元素的换行符
            split_str = line.split(': ')
            key = split_str[0]
            value = split_str[1]
            if replace_key != None:
                index = replace_key.index(key) if key in replace_key else -1
                if index != -1:
                    value = replace_value[index]
            data[key] = value

    # 有替换的内容，需要确认一下
    if replace_key != None and len(replace_key) >0:
        print('headers:')
        for i in range(len(replace_key)):
            print(f'{replace_key[i]}:{replace_value[i]}')
        print()
    return data

def request_http(url,headers,payload=None,method='POST'):
    global proxy_pool
    global proxy_cycle
    try:
        if payload == None:
            response = session.request(method, url, headers=headers, timeout=15)
        else:
            response = session.request(method, url, headers=headers, data=json.dumps(payload), timeout=15)
        # {"isHolderOfCollection":true}
        data = response.json()
        # print('获取是否持有')
        # print('返回数据：',data)
        return data
    except Exception as ex:
        print('出现错误：%s' % ex)
        if (isinstance(ex.args[0], str) or isinstance(ex.args[0], tuple)) and ex.args[0].find('timed out') != -1:
            print(session.proxies)
            print(proxy_pool)
            # proxy_cycle = itertools.cycle(proxy_pool)
        # traceback.print_exc()
        global need_mint
        need_mint = True




def claim(addr,pk):
    global is_batch
    try:

        gasPrice = int((w3.toWei(chain_gas_price,'gwei') if chain_gas_price != 0 else w3.eth.gas_price) * default_gas_price)
        params = {
            'gas':gas_limit,
            'gasPrice':gasPrice,
            'nonce':w3.eth.getTransactionCount(addr),
            'value':0,
            'chainId':chainId,
        }
        # mint(uint256 supplyType, uint256 maxAmount, uint8 v, bytes32 r, bytes32 s)
        # 直接提交请求
        supplyType = 0
        maxAmount = 2646365509777468087
        func = mintContract.functions.mint(supplyType,maxAmount)
        # 直接提交请求
        tx = func.buildTransaction(params)
        sign_tx = w3.eth.account.sign_transaction(tx, private_key=pk)
        r = sign_tx['r']
        s = sign_tx['s']
        v = sign_tx['v']
        func = mintContract.functions.mint(supplyType, maxAmount,v,r,s)

        # 直接提交请求
        tx = func.buildTransaction(params)
        sign_tx = w3.eth.account.sign_transaction(tx, private_key=pk)
        print('提交请求:{0}'.format(sign_tx))
        tx_hash = w3.eth.sendRawTransaction(sign_tx.rawTransaction)
        if is_batch:
            return tx_hash
        else:
            receipt = searchReceipt(tx_hash)
            if receipt['status'] == 1:
                print('Mint 成功')
                global gas_expend
                gas_expend += w3.fromWei(receipt['effectiveGasPrice'] * receipt['gasUsed'], 'ether')
                # 开启批量
                is_batch = True
                hash_list_and_num[tx_hash] = num
                print('开启批量模式')
                return True
            else:
                print('Mint 失败， {0}{1}'.format(block_chain_explorer,w3.toHex(tx_hash)))
                return False
    except Exception as ex:
        print('出现错误：%s' % ex)
        return False


def claim_data(addr,pk,signature):
    global is_batch
    try:
        data = f'0xf75e0384000000000000000000000000000000000000000000000000000000000000002' \
               f'0000000000000000000000000{addr[2:]}' \
               f'0000000000000000000000000000000000000000000000000000000000000001' \
               f'0000000000000000000000000000000000000000000000000000000000000060' \
               f'0000000000000000000000000000000000000000000000000000000000000041' \
               f'{signature[2:]}' \
               f'00000000000000000000000000000000000000000000000000000000000000'

        gasPrice = int(w3.toWei(1,'gwei') * default_gas_price)
        params = {
            'from': addr,
            'to': w3.to_checksum_address('0xc92df682a8dc28717c92d7b5832376e6ac15a90d'),
            'gas':gas_limit,
            'gasPrice':gasPrice,
            'nonce':w3.eth.getTransactionCount(addr),
            'value':0,
            'chainId':chainId,
            'data':data
        }
        # mint(address to,uint256 amount,uint256 max_amount,bytes signature_)
        # 直接提交请求
        sign_tx = w3.eth.account.sign_transaction(params, private_key=pk)
        print('提交请求:{0}'.format(sign_tx))
        tx_hash = w3.eth.sendRawTransaction(sign_tx.rawTransaction)
        if is_batch:
            return tx_hash
        else:
            receipt = searchReceipt(tx_hash)
            if receipt['status'] == 0:
                print('Mint 失败， {0}{1}'.format(block_chain_explorer, w3.toHex(tx_hash)))
                return False
            else:
                print('Mint 成功')
                global gas_expend
                gas_expend += w3.fromWei(receipt['effectiveGasPrice'] * receipt['gasUsed'], 'ether')
                # 开启批量
                is_batch = True
                hash_list_and_num[tx_hash] = num
                print('开启批量模式')
                return True
    except Exception as ex:
        print('出现错误：%s' % ex)
        return False

# 今日-本周签到时间=本周可签到天数
def can_sign_days(week_sing_days):
    today = datetime.strptime(get_today_time(), '%Y-%m-%d')
    # 本周签到日期
    week_date = datetime.strptime(week_sing_days, '%Y-%m-%d')
    delta = today - week_date
    # 当天也要算
    return delta.days + 1


# 不确定是几点更新签到，默认是早上8点，没到8点之前就算前一天
def get_today_time():
    delta = timedelta(hours=-8)
    now = datetime.now()
    live_time = now + delta
    return_time = live_time.strftime('%Y-%m-%d')
    # print('secondlive时间：{0}'.format(live_time.strftime('%Y-%m-%d %H:%M:%S')))
    return return_time

#
# # 获得预签名明文
# def get_pre_sign_text(addr,session):
#     # {"nonce":"NrJv8IVPwMWOKUHJS","address":"0x84B35Fdc0A8EA0112Bb8BBBD1f3CCFBbAbACBd93","expires_at":"2023-05-12T09:47:58.104Z"}
#     url = 'https://auth.privy.io/api/v1/siwe/init'
#     payload = {
#       "address": addr
#     }
#     headers = {
#         'accept': 'application/json, text/plain, */*',
#         'accept-encoding': 'gzip, deflate, br',
#         'accept-language': 'zh-CN,zh;q=0.9',
#         'content-length': '56',
#         'content-type': 'application/json',
#         'origin': 'https://optimism.decent.xyz',
#         'privy-app-id': 'clcqleb81000ckz08oxqo6t9t',
#         'privy-client': 'react-auth:1.23.0',
#         'referer': 'https://optimism.decent.xyz/',
#         'sec-ch-ua': '"Chromium";v="112", "Google Chrome";v="112", "Not:A-Brand";v="99"',
#         'sec-ch-ua-mobile': '?0',
#         'sec-ch-ua-platform': '"Windows"',
#         'sec-fetch-dest': 'empty',
#         'sec-fetch-mode': 'cors',
#         'sec-fetch-site': 'cross-site',
#         'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/112.0.0.0 Safari/537.36'
#     }
#     while True:
#         try:
#             response = session.request("POST", url, headers=headers, data=json.dumps(payload), timeout=20)
#             data = response.json()
#             data['nonce']
#             # print(data)
#             print('获取预签名信息')
#             return data
#         except Exception as e:
#             print('获取预签名信息错误：{0}'.format(e))
#             continue

# 默认签名方式
def get_pre_sign_data(srcdata,pk):
    msg = srcdata
    msghash = defunct_hash_message(text=msg)
    key = pk
    text=w3.eth.account.signHash(msghash, key)
    sign_data=text.signature.hex()
    print('执行签名')
    return (sign_data)

# 把ip截取出来
def pause_ip_by_proxy(proxy: str):
    if proxy.find('@') != -1:
        ip = proxy.split('@')[1].split(':')[0]
    else:
        ip = proxy.split('//')[1].split(':')[0]
    return ip

ip_search_model = 1

# 用代理创建一个session对象
def create_session(proxy: str) -> requests.Session:
    session = requests.Session()
    session.proxies = {"http": proxy, "https": proxy}
    global ip_search_model
    try:
        if ip_search_model == 1:
            resp = session.get('http://ip-api.com/json/', timeout=10)
            data = resp.json()
            ip = data['query']
            country = data['country']
            city = data['city']
            print('代理ip：{0}，地区：{1}->{2}'.format(ip, country,city))
        else:
            ip = pause_ip_by_proxy(proxy)
            if ip != '127.0.0.1':
                resp = session.get('https://opendata.baidu.com/api.php?query={0}&co=&resource_id=6006&oe=utf8'.format(ip))
                data = resp.json()['data'][0]
                ip = data['OriginQuery']
                country = data['location']
                print('代理ip：{0}，地区：{1}'.format(ip,country))
                ip_search_model = 2
    except:
        ip_search_model = 2
        return None

    retries = Retry(total=5, backoff_factor=0.1, status_forcelist=[500, 502, 503, 504])
    session.mount("http://", HTTPAdapter(max_retries=retries))
    session.mount("https://", HTTPAdapter(max_retries=retries))
    return session



def get_config():
    config_file = open('config.json', 'r',encoding='utf-8')
    config_text = config_file.read()
    config_json = json.loads(config_text)
    return config_json[0]

def init_proxy_pool():
    # 获取ip地址
    excel_data = pd.read_excel('ip.xlsx')
    ip_list=excel_data['ip'].to_list()
    port_list=excel_data['port'].to_list()
    username_list=excel_data['username'].to_list()
    password_list=excel_data['password'].to_list()
    proxy_list = []
    for i in range(len(ip_list)):
        proxy_str = 'http://{0}:{1}@{2}:{3}'.format(username_list[i],password_list[i],ip_list[i],port_list[i])
        proxy_list.append(proxy_str)

    return proxy_list

def init_excel_header(file_name,column_arr):
    # 打开Excel文件
    workbook = openpyxl.load_workbook(file_name)
    # 选择第一个工作表
    worksheet = workbook.active

    for column_name in column_arr:
        # 检查每一列是否包含表头
        header_exists = False
        for col in worksheet.iter_cols(min_row=1, max_row=1):
            for cell in col:
                if cell.value == column_name:
                    header_exists = True
                    break

        # 如果表头不存在，则在最后一列添加表头
        if not header_exists:
            # 获取最后一列的列字母
            last_column_index = worksheet.max_column
            last_column_letter = openpyxl.utils.get_column_letter(last_column_index + 1)
            # 在最后一列添加表头
            worksheet[f"{last_column_letter}1"] = column_name
            print('自动添加表头：{0}'.format(column_name))

    # 保存Excel文件
    workbook.save(file_name)

def get_next_sign_time():
    now = datetime.now()
    days = 0
    hours = 8
    minutes = 0
    if now.hour >= 8:
        # 超过当天下午4点，天数+1
        days = 1
    default_time = datetime(year=now.year, month=now.month, day=now.day) + timedelta(days=days, hours=hours,minutes=minutes)
    return str(default_time)

# 是否可以签到
def can_sign(next_time):
    if isinstance(next_time,float):
        return True

    if isinstance(next_time,str) and str == '':
        return True

    next_time = datetime.strptime(next_time, "%Y-%m-%d %H:%M:%S")
    # 领取时间小于当前时间，就是可领
    can_punch = next_time < datetime.now()
    if can_punch:
        print('签到时间已到：{0}'.format(next_time))
    else:
        print('签到时间未到：{0}'.format(next_time))
    return can_punch

# 返回access_token,邀请码
def get_access_token(addr, pk ,user_agent, invitationCode):
    url = "https://alienswap.xyz/alien-api/api/v1/public/user/signin?network=eth"
    headers = {
        'Accept': 'application/json, text/plain, */*',
        'User-Agent': user_agent
    }

    sign_text = "Welcome to AlienSwap!\nClick to sign in and accept the AlienSwap Terms of Service.\nThis request will not trigger a blockchain transaction or cost any gas fees."
    sign = get_pre_sign_data(sign_text, pk)

    payload = {
        "address": addr,
        "signature": sign,
        "nonce": sign_text,
        "inviter": invitationCode,
        "src": 4,
        "network": "eth"
    }

    # {
    #     "status": 0,
    #     "msg": "ok",
    #     "data": {
    #         "is_new": true,
    #         "invite_code": "rhiMcB",
    #         "access_token": "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpYXQiOjE2OTQwNjg4ODQsInN1YiI6OTI1NzU1LCJ0eXBlIjoiYWNjZXNzIiwiZXhwIjoxNjk1MzY0ODg0fQ.Dn6FG8kFB9hqodNqerPUn6XFKjyPf1CDuqZbHdrKEEY"
    #     }
    # }
    data = request_http(url, headers=headers, payload=payload)

    current_invite_code = data["data"]["invite_code"]
    current_access_token = data["data"]["access_token"]
    print(f'签名成功，邀请码：{current_invite_code}')
    return (current_access_token, current_invite_code)

proxy_pool = []
proxy_pool_model = False

# 开始
start_time = time.time()
print('开始Mint任务==')

config = get_config()
dataScope = config['dataScope']
httpProxy = config['httpProxy']
if httpProxy == '':
    httpProxy = 'http://user293:7W1EIMLI@154.212.178.190:43147'
print('配置文件信息：')
try:
    excel_data = pd.read_excel('ip.xlsx')
    print('检测到ip.xlsx文件，走代理池模式')
    proxy_pool = init_proxy_pool()
    proxy_cycle = itertools.cycle(proxy_pool)
    proxy_pool_model = True
except:
    print('HTTP代理：{0}'.format(config['httpProxy']))
    print('HTTPS代理：{0}'.format(config['httpsProxy']))

try:
    config_gas_price = config['defaultGasPriceMultiple']
    if isinstance(config_gas_price,int) or isinstance(config_gas_price,float):
        default_gas_price = config_gas_price
    else:
        default_gas_price = 1
except:
    default_gas_price = 1

is_batch = False

session = None
while session is None:
    # print('代理ip：{0}'.format(proxy.split('@')[1].split(':')[0]))
    if proxy_pool_model:
        httpProxy = next(proxy_cycle)
    session = create_session(httpProxy)

w3 = Web3(Web3.HTTPProvider(rpc_url,session=session))
if block_chain.middleware == 'poa':
    w3.middleware_onion.inject(geth_poa_middleware, layer=0)

# gas不足的地址
gas_not_enough_add = []
DEFAULT_ABI = '[{"inputs":[],"name":"name","outputs":[{"internalType":"string","name":"","type":"string"}],"stateMutability":"view","type":"function"},{"inputs":[],"name":"symbol","outputs":[{"internalType":"string","name":"","type":"string"}],"stateMutability":"view","type":"function"},{"inputs":[],"name":"decimals","outputs":[{"internalType":"uint8","name":"","type":"uint8"}],"stateMutability":"view","type":"function"},{"inputs":[],"name":"totalSupply","outputs":[{"internalType":"uint256","name":"","type":"uint256"}],"stateMutability":"view","type":"function"},{"inputs":[{"internalType":"address","name":"account","type":"address"}],"name":"balanceOf","outputs":[{"internalType":"uint256","name":"","type":"uint256"}],"stateMutability":"view","type":"function"},{"inputs":[{"internalType":"address","name":"recipient","type":"address"},{"internalType":"uint256","name":"amount","type":"uint256"}],"name":"transfer","outputs":[{"internalType":"bool","name":"","type":"bool"}],"stateMutability":"nonpayable","type":"function"},{"inputs":[{"internalType":"address","name":"sender","type":"address"},{"internalType":"address","name":"recipient","type":"address"},{"internalType":"uint256","name":"amount","type":"uint256"}],"name":"transferFrom","outputs":[{"internalType":"bool","name":"","type":"bool"}],"stateMutability":"nonpayable","type":"function"},{"inputs":[{"internalType":"address","name":"spender","type":"address"},{"internalType":"uint256","name":"amount","type":"uint256"}],"name":"approve","outputs":[{"internalType":"bool","name":"","type":"bool"}],"stateMutability":"nonpayable","type":"function"},{"inputs":[{"internalType":"address","name":"owner","type":"address"},{"internalType":"address","name":"spender","type":"address"}],"name":"allowance","outputs":[{"internalType":"uint256","name":"","type":"uint256"}],"stateMutability":"view","type":"function"},{"anonymous":false,"inputs":[{"indexed":true,"internalType":"address","name":"from","type":"address"},{"indexed":true,"internalType":"address","name":"to","type":"address"},{"indexed":false,"internalType":"uint256","name":"value","type":"uint256"}],"name":"Transfer","type":"event"},{"anonymous":false,"inputs":[{"indexed":true,"internalType":"address","name":"owner","type":"address"},{"indexed":true,"internalType":"address","name":"spender","type":"address"},{"indexed":false,"internalType":"uint256","name":"value","type":"uint256"}],"name":"Approval","type":"event"}]'
success_count = 0



# TODO 要修改为mint的合约地址
# mint合约
mintContractAddr = w3.toChecksumAddress('0xc92df682a8dc28717c92d7b5832376e6ac15a90d')
MINT_ABI = DEFAULT_ABI
mintContract = w3.eth.contract(address=mintContractAddr,abi=MINT_ABI)


# TODO 要修改为标志mint状态的表头名字
# 用于标志的excel表头名字
excel_file_name = 'eth.xlsx'
mint_status = 'Alien2049Adventure'
init_excel_header(excel_file_name, [mint_status, 'AlienAccessToken'])

# TODO 修改为交互需要的gasLimit
gas_limit = 192777


need_mint = True
while need_mint:
    need_mint = False
    excel_data = pd.read_excel(excel_file_name)
    # 获取助记词并转换为列表格式
    Address = excel_data['Address'].to_list()
    PrivateKey = excel_data['PrivateKey'].to_list()
    Minted = excel_data[mint_status].to_list()
    AlienAccessToken = excel_data['AlienAccessToken'].to_list()

    # 获取钱包私钥

    if dataScope:
        maxSize = len(PrivateKey)
        configSize = dataScope[1]
        if maxSize < configSize:
            configSize = maxSize

        Address = Address[dataScope[0] - 1:configSize]
        PrivateKey = PrivateKey[dataScope[0] - 1:configSize]
        Minted = Minted[dataScope[0] - 1:configSize]
        AlienAccessToken = AlienAccessToken[dataScope[0] - 1:configSize]

    total = len(PrivateKey)
    print("地址数： {}".format(total))
    error_address = []
    gas_not_enough_add = []
    hash_list_and_num = {}
    invitationCode = "rhiMcB"
    # invitationCodesIndex = 0
    # invitationCount = 0

    for i, priv in enumerate(PrivateKey):
        gas_expend = Decimal('0')
        excel_modify = False
        try:
            pk = priv
            subAccount = Account.from_key(pk)
            addr = w3.toChecksumAddress(subAccount.address)
            print('{0}/{1} {2}'.format(i+dataScope[0], total+dataScope[0]-1, addr))

            print(f'当前使用邀请码：{invitationCode}')

            if isinstance(Minted[i], str) and Minted[i] == 'YES':
                print('已经Mint过')
                print('----------------------------------------\n')
                continue

            num = i + dataScope[0] - 1

            if proxy_pool_model:
                session = None
                while session is None:
                    proxy = next(proxy_cycle)
                    # print('代理ip：{0}'.format(proxy.split('@')[1].split(':')[0]))
                    session = create_session(proxy)
                w3 = Web3(Web3.HTTPProvider(rpc_url, session=session))
                if block_chain.middleware == 'poa':
                    w3.middleware_onion.inject(geth_poa_middleware, layer=0)

            ua = UserAgent()
            user_agent = ua.random

            access_token = AlienAccessToken[i]
            # 如果没有值，默认是float
            if not isinstance(access_token, str):
                access_token = ''

            while access_token == '':
                access_token, invitationCode = get_access_token(addr, pk, user_agent, invitationCode)
                # 更新到文档
                if access_token != '':
                    excel_data.loc[num:num, 'AlienAccessToken'] = access_token


            headers = {
                'Accept': 'application/json, text/plain, */*',
                'User-Agent': user_agent,
                'Authorization': 'Bearer ' + access_token
            }


            # mintNft
            url = "https://alienswap.xyz/alien-api/api/v1/private/activity/token2049/account/mint?network=eth"
            payload = {
              "partner_uuid": "gDgt2BBJ785S2zqtv5aREn",
              "network": "eth"
            }
            # {
            #     "status": 0,
            #     "msg": "ok",
            #     "data": {
            #         "token_id": 9059,
            #         "image_url": "https://alien-static.s3.ap-northeast-1.amazonaws.com/files/edebdfc8fc2204929e34d30c72743681.png"
            #     }
            # }
            data = request_http(url, headers=headers, payload=payload)
            if data['status'] == 401 and 'msg' in data and data['msg'] == 'Unauthorized':
                print('签名过期，重新获取')
                access_token, invitationCode = get_access_token(addr, pk, user_agent, invitationCode)
                # 更新到文档
                if access_token != '':
                    excel_data.loc[num:num, 'AlienAccessToken'] = access_token
                    headers.update({
                        'Authorization': 'Bearer ' + access_token
                    })
                    data = request_http(url, headers=headers, payload=payload)

            if data['status'] == 2707:
                print(f'已经Mint过')
                excel_data.loc[num:num, mint_status] = 'YES'
            else:
                token_id = data['data']['token_id']
                print(f'Mint成功，TokenId：{token_id}')
                success_count += 1
                excel_data.loc[num:num, mint_status] = 'YES'

            excel_data.to_excel(excel_file_name, index=False, engine="openpyxl")
            print('----------------------------------------\n')
        except Exception as ex:
            excel_data.to_excel(excel_file_name, index=False, engine="openpyxl")
            print('出现错误：%s' % ex)
            if isinstance(ex.args[0], str) and ex.args[0].find('Too Many Requests') != -1:
                print('休息5秒')
                time.sleep(5)
            # traceback.print_exc()
            need_mint = True
            print('----------------------------------------\n')
            continue

    # 查询最后一批
    if is_batch and len(hash_list_and_num)>0:
        batchSerachReceiptAndSaveExcel(hash_list_and_num)

print('成功Mint账号数：{0}'.format(success_count))
if len(gas_not_enough_add)>0:
    print('Gas不足的钱包：{0}'.format(gas_not_enough_add))

end_time = time.time()
duration = end_time-start_time
hours, rem = divmod(duration, 3600)
minutes, seconds = divmod(rem, 60)
print("用时: {:0>2}:{:0>2}:{:0>2}".format(int(hours), int(minutes), int(seconds)))

print('\n\n===========================================')
user_password = getpass.getpass(prompt='程序执行完毕，请关闭！')

